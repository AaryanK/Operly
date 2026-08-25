from __future__ import annotations

from io import BytesIO
from types import SimpleNamespace

import pytest
import pytest_asyncio
from openpyxl import load_workbook
from PIL import Image
from sqlalchemy.ext.asyncio import create_async_engine

from packages.artifacts.service import ArtifactScope, ArtifactService
from packages.capabilities.file_authoring_provider import FileAuthoringProvider
from packages.database.artifact_models import ArtifactRecord
from packages.database.db import Base
from packages.database.models import Tenant


@pytest_asyncio.fixture
async def runtime_db():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as connection:
        await connection.run_sync(
            lambda sync: Base.metadata.create_all(
                sync,
                tables=[Tenant.__table__, ArtifactRecord.__table__],
            )
        )
    async with engine.connect() as connection:
        transaction = await connection.begin()
        from sqlalchemy.ext.asyncio import AsyncSession

        db = AsyncSession(bind=connection, expire_on_commit=False)
        db.add(Tenant(id="tenant-a", name="Tenant A", slug="tenant-a"))
        await db.flush()
        yield db
        await db.close()
        await transaction.rollback()
    await engine.dispose()


def context(db):
    return SimpleNamespace(
        tenant_id="tenant-a",
        actor_id="user-1",
        scope_kind="workspace",
        scope_id="tenant-a",
        owner_user_id=None,
        execution_id="call-1",
        db=db,
        invocation={"metadata": {"runtime_run_id": "run-1"}},
    )


@pytest.mark.asyncio
async def test_create_spreadsheet_preserves_declared_schema(runtime_db):
    provider = FileAuthoringProvider()
    ctx = context(runtime_db)
    columns = ["name", "email", "classification", "recommended next action"]
    result = await provider.execute(
        ctx,
        "files.create_spreadsheet",
        {
            "title": "Contact Engagement",
            "filename": "contact-engagement.xlsx",
            "output_format": "xlsx",
            "columns": columns,
            "rows": [
                {
                    "name": "Example Person",
                    "email": "example@example.com",
                    "classification": "Email engaged",
                    "recommended next action": "Follow up",
                }
            ],
        },
    )
    verified = await provider.verify(ctx, "files.create_spreadsheet", {}, result)
    assert verified.success is True
    assert verified.evidence["columns"] == columns
    assert verified.evidence["row_count"] == 1

    artifact_id = verified.evidence["artifact_ids"][0]
    raw = await ArtifactService(runtime_db).read_bytes(
        ArtifactScope("workspace", "tenant-a", tenant_id="tenant-a"), artifact_id
    )
    workbook = load_workbook(BytesIO(raw), read_only=True)
    worksheet = workbook.active
    material = list(worksheet.iter_rows(values_only=True))
    assert list(material[0]) == columns
    assert list(material[1]) == [
        "Example Person",
        "example@example.com",
        "Email engaged",
        "Follow up",
    ]
    assert "attachment_index" not in material[0]


@pytest.mark.asyncio
async def test_create_document_generates_real_pdf_not_attachment_report(runtime_db):
    provider = FileAuthoringProvider()
    ctx = context(runtime_db)
    result = await provider.execute(
        ctx,
        "files.create_document",
        {
            "title": "Executive Summary",
            "content": "# Summary\n- Four contacts reviewed\n- One follow-up required",
            "output_format": "pdf",
            "filename": "executive-summary.pdf",
        },
    )
    verified = await provider.verify(ctx, "files.create_document", {}, result)
    assert verified.success is True
    assert verified.evidence["artifact_kind"] == "document"
    assert verified.evidence["output_format"] == "pdf"

    raw = await ArtifactService(runtime_db).read_bytes(
        ArtifactScope("workspace", "tenant-a", tenant_id="tenant-a"),
        verified.evidence["artifact_ids"][0],
    )
    assert raw.startswith(b"%PDF")


@pytest.mark.asyncio
async def test_image_to_pdf_conversion_is_deterministic(runtime_db):
    ctx = context(runtime_db)
    service = ArtifactService(runtime_db)
    image = Image.new("RGB", (20, 10), "white")
    source_bytes = BytesIO()
    image.save(source_bytes, format="PNG")
    source = await service.create_bytes(
        ArtifactScope("workspace", "tenant-a", tenant_id="tenant-a"),
        filename="source.png",
        content_type="image/png",
        content=source_bytes.getvalue(),
        source="test",
    )

    provider = FileAuthoringProvider()
    result = await provider.execute(
        ctx,
        "files.convert",
        {"artifact_id": source.id, "output_format": "pdf", "filename": "source.pdf"},
    )
    verified = await provider.verify(ctx, "files.convert", {}, result)
    assert verified.success is True
    assert verified.evidence["conversion"] == "image_to_pdf"
    raw = await service.read_bytes(
        ArtifactScope("workspace", "tenant-a", tenant_id="tenant-a"),
        verified.evidence["artifact_ids"][0],
    )
    assert raw.startswith(b"%PDF")


@pytest.mark.asyncio
async def test_csv_to_xlsx_preserves_source_columns(runtime_db):
    ctx = context(runtime_db)
    service = ArtifactService(runtime_db)
    source = await service.create_bytes(
        ArtifactScope("workspace", "tenant-a", tenant_id="tenant-a"),
        filename="people.csv",
        content_type="text/csv",
        content=b"person,email,classification\nA,a@example.com,Email only\n",
        source="test",
    )
    provider = FileAuthoringProvider()
    result = await provider.execute(
        ctx,
        "files.convert",
        {"artifact_id": source.id, "output_format": "xlsx", "filename": "people.xlsx"},
    )
    verified = await provider.verify(ctx, "files.convert", {}, result)
    assert verified.success is True
    assert verified.evidence["columns"] == ["person", "email", "classification"]
    assert verified.evidence["row_count"] == 1
